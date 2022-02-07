import datetime as dt
import pandas as pd
from pandas_datareader import data as pdr
import yfinance as yf
import xlsxwriter
import numpy as np
from openpyxl import load_workbook
yf.pdr_override()


start ='2020-01-02'
now = dt.datetime.now().date()

allstocks = 'RichardStocks.xlsx'
output = 'VCPlog.xlsx'

stocklist = pd.read_excel(f"/Users/benjaminluk/PycharmProjects/VCPscreener/{allstocks}")  #change this

# Intermediate files
Template_Screen_file = 'VCP_template_screen.xlsx'
RS_Ranked_File = 'RS_Ranked.xlsx'
Output_file = f'/Users/benjaminluk/PycharmProjects/VCPscreener/{output}' #change this



# Parameters
RS_percentile = 0.3 #top 30% of RS Rating wanted
delta = 3 # +-3 days from minimum or maximum point to find lowest/highest close
peakupperbound = 1.02 #upper bound of all other peaks relative to first peak
peaklowerbound = 0.95 #lower bound of all other peaks relative to first peak
ddleniency = 0.30 #each max to min length is at least 0.3 of previous



exportList = []
exportList.append(['Stock', "RS_Rating", "50 Day MA", "150 Day MA", "200 Day MA", "52 Week Low", "52 week High"])

for i in stocklist.index:
    stock = str(stocklist["Symbol"][i])
    try:
        df = pdr.get_data_yahoo(stock, start, now)

        smaUsed = [50, 150, 200]
        for x in smaUsed:
            sma = x
            df["SMA_" + str(sma)] = round(df.iloc[:, 4].rolling(window=sma).mean(), 2)

        currentClose = df["Adj Close"][-1]
        moving_average_50 = df["SMA_50"][-1]
        moving_average_150 = df["SMA_150"][-1]
        moving_average_200 = df["SMA_200"][-1]
        low_of_52week = min(df["Adj Close"][-260:])
        high_of_52week = max(df["Adj Close"][-260:])


        threemonthclose = df['Adj Close'][-63]
        sixmonthclose = df['Adj Close'][-126]
        ninemonthclose = df['Adj Close'][-189]
        twelvemonthclose = df['Adj Close'][-252]


        RS_Rating = (threemonthclose / currentClose) * 2 + (sixmonthclose / currentClose) + (ninemonthclose / currentClose) + (twelvemonthclose / currentClose)


        try:
            moving_average_200_20past = df["SMA_200"][-20]
        except Exception:
            moving_average_200_20past = 0

        print("Checking " + stock + ".....")

        # Condition 1: Current Price > 150 SMA and > 200 SMA
        if (currentClose > moving_average_150 and currentClose > moving_average_200):
            cond1 = True
        else:
            cond1 = False
        # Condition 2: 150 SMA and > 200 SMA
        if (moving_average_150 > moving_average_200):
            cond2 = True
        else:
            cond2 = False
        # Condition 3: 200 SMA trending up for at least 1 month (ideally 4-5 months)
        if (moving_average_200 > moving_average_200_20past):
            cond3 = True
        else:
            cond3 = False
        # Condition 4: 50 SMA> 150 SMA and 50 SMA> 200 SMA
        if (moving_average_50 > moving_average_150 and moving_average_50 > moving_average_200):
            cond4 = True
        else:
            cond4 = False
        # Condition 5: Current Price > 50 SMA
        if (currentClose > moving_average_50):
            cond5 = True
        else:
            cond5 = False
        # Condition 6: Current Price is at least 30% above 52 week low (Many of the best are up 100-300% before coming out of consolidation)
        if (currentClose >= 1.30 * low_of_52week):
            cond6 = True
        else:
            cond6 = False
        # Condition 7: Current Price is within 25% of 52 week high
        if (currentClose >= 0.75 * high_of_52week):
            cond7 = True
        else:
            cond7 = False


        if (cond1 and cond2 and cond3 and cond4 and cond5 and cond6 and cond7):
            exportList.append([stock, RS_Rating, moving_average_50,
                                            moving_average_150, moving_average_200,
                                            low_of_52week, high_of_52week])


    except Exception:
        print("No data on " + stock)




with xlsxwriter.Workbook(Template_Screen_file) as workbook:
    worksheet = workbook.add_worksheet()

    for row_num, data in enumerate(exportList):
        worksheet.write_row(row_num, 0, data)


#=================================================#

#Sort by RS_Rating, want top 30%
xl = pd.ExcelFile(Template_Screen_file)
df = xl.parse("Sheet1")
df = df.sort_values(by=["RS_Rating"], ascending = False)
lengthneeded = int(len(df) * RS_percentile)
df = df[: lengthneeded]

writer = pd.ExcelWriter(RS_Ranked_File)
df.to_excel(writer,sheet_name='Sheet1',index=False)
writer.save()

#================================================#

##VCP Detection Main Code

def localmaxmin(datamaxmin):
    min_max = np.diff(np.sign(np.diff(datamaxmin))).nonzero()[0] + 1  # local min & max
    l_min = (np.diff(np.sign(np.diff(datamaxmin))) > 0).nonzero()[0] + 1  # local min
    l_max = (np.diff(np.sign(np.diff(datamaxmin))) < 0).nonzero()[0] + 1  # local max
    # +1 due to the fact that diff reduces the original index number

    return l_min, l_max


def maxminrange(datarange, y_pol, delta):
    # extend the suspected x range:
    # how many ticks to the left and to the right from local minimum on x axis

    dict_min = dict()
    dict_max = dict()
    dict_x = dict()

    df_len = len(datarange)  # number of rows in dataset

    l_min = localmaxmin(y_pol)[0]
    l_max = localmaxmin(y_pol)[1]

    # l_min value ranges
    for element in l_min:  # x coordinates of suspected minimums
        l_bound = element - delta  # lower bound (left)
        u_bound = element + delta  # upper bound (right)
        x_range = range(l_bound, u_bound + 1)  # range of x positions where we SUSPECT to find a low
        dict_min[
            element] = x_range  # just helpful dictionary that holds suspected x ranges for further visualization strips

        min_loc_list = list()
        for x_element in x_range:
            if x_element > 0 and x_element < df_len:  # need to stay within the dataframe
                # y_loc_list.append(ticker_df.Low.iloc[x_element])   # list of suspected y values that can be a minimum
                min_loc_list.append(datarange['Low'].iloc[x_element])
                # print(y_loc_list)
                # print('ticker_df.Low.iloc[x_element]', ticker_df.Low.iloc[x_element])
        dict_min[element] = min_loc_list  # key in element is suspected x position of minimum
        # to each suspected minimums we append the price values around that x position
        # so 40: [53.70000076293945, 53.93000030517578, 52.84000015258789, 53.290000915527344]
        # x position: [ 40$, 39$, 41$, 45$]

    for element in l_max:  # x coordinates of suspected minimums
        l_bound = element - delta  # lower bound (left)
        u_bound = element + delta  # upper bound (right)
        x_range = range(l_bound, u_bound + 1)  # range of x positions where we SUSPECT to find a low
        dict_max[
            element] = x_range  # just helpful dictionary that holds suspected x ranges for further visualization strips

        max_loc_list = list()
        for y_element in x_range:
            if y_element > 0 and y_element < df_len:  # need to stay within the dataframe

                max_loc_list.append(datarange['High'].iloc[y_element])
                # print(y_loc_list)
                # print('ticker_df.Low.iloc[x_element]', ticker_df.Low.iloc[x_element])
        dict_max[element] = max_loc_list

    return dict_min, dict_max


def VCP_finder(datafind, delta, peakupperbound, peaklowerbound, ddleniency):

    ok = []
    windowdeg = [[40, 11], [60, 11], [80, 11], [100, 9], [120, 9]]
    currentWindow = [0, 0, 0, 0, 0]

    # data splicing for rolling windows
    buysignal = False

    for i in np.arange(0, len(windowdeg) - 1, 1):
        window = windowdeg[i][0]
        polydeg = windowdeg[i][1]

        datanow = datafind[-window + 1:]
        predata = datafind[-2 * window + 1: -window]

        # fit polynomial curve through data
        # choose the input and output variables
        x, y = np.arange(0, window - 1, 1), datanow['Adj Close']
        # curve fit
        popt = np.polyfit(x, y, polydeg)
        y_pol = np.polyval(popt, x)

        # minimum and maximum ranges
        dict_min = maxminrange(datanow, y_pol, delta)[0]
        dict_max = maxminrange(datanow, y_pol, delta)[1]

        min_loclist = localmaxmin(y_pol)[0]
        max_loclist = localmaxmin(y_pol)[1]

        for k in min_loclist:
            dict_min[k] = np.min(dict_min[k])

        min_list = list(dict_min.values())
        min_values = np.array(min_list)

        for j in max_loclist:
            dict_max[j] = np.max(dict_max[j])

        max_list = list(dict_max.values())
        max_values = np.array(max_list)

        ###VCP conditions
        conditions = np.zeros(5)

        ##1. Stage 2 growth present/ Past uptrend
        xs, ys = np.arange(0, window - 1, 1), predata['Adj Close']
        prepopt = np.polyfit(xs, ys, 1)
        if prepopt[0] > 0:
            conditions[0] = True

        ##2. wedgeup/ triangle ascending
        if (len(min_values) > len(max_values)):  # make sure graph starts with a peak
            min_values = min_values[1:]
        elif (len(min_values) < len(max_values)):
            max_values = max_values[:-1]

        if ((min_values + max_values).size != 0):
            if (len(min_list) == len(max_list)) and ((min_loclist[0] < max_loclist[0])):
                max_values = max_values[:-1]                # making sure the window has pattern as follows:
                min_values = min_values[1:]                 #          . . .
                                                            #           . . .          (max first followed by min, and end with min)


        if len(min_values) > 1:
            troughline = np.polyfit(np.arange(0, len(min_values), 1), min_values, 1)[0]    #fit a linear line through minimums
        else:
            troughline = 0


        if troughline > 0:  # ascending troughs, check gradient > 0
            conditions[1] = True
        else:
            conditions[1] = False

        ##3. first peak greater than all other peaks (with leniency parameter)
        for p in max_values:
            if p < (max_values[0] * peakupperbound) and p > (max_values[0] * peaklowerbound):
                conditions[2] = True
            else:
                conditions[2] = False

        if len(max_values) <= 1:  # make sure at least two peaks
            conditions[2] = False

        ##4. drawdown contracting
        drawdowns = np.abs(max_values - min_values)
        conditions[3] = all((ddleniency) * earlier <= later for earlier, later in zip(drawdowns, drawdowns[1:])) #check each max min length is shortening


        ##5. Decreasing Volume in last peak
        if min_loclist.size == 0 and max_loclist.size == 0:  # ignore empty minimum and maximum lists
            min_loclist = np.array([0])
            max_loclist = np.array([0])
        if max_loclist.size == 0:
            max_loclist = min_loclist
        if min_loclist.size == 0:
            min_loclist = max_loclist

        lastpeak = max_loclist[-1]
        lasttrough = min_loclist[-1]

        volumes = datanow['Volume'][lastpeak:lasttrough]

        volx = np.arange(0, len(volumes), 1)

        if len(volumes) > 1:
            volpopt = np.polyfit(volx, volumes, 1)[0]
        else:
            volpopt = 0

        if volpopt < 0:
            conditions[4] = True

        ## Optional Condition: first wave has > 8% dd
        # if drawdowns.size != 0:
        #    if (drawdowns[0]/max_values[0]) > 0.10:
        #        conditions[5] = True

        # print ones where all conditions are True
        if all(conditions) == True:
            currentWindow[i] = True
            buysignal = True


    return buysignal, currentWindow

####################################################

stocklist = df
symbols = stocklist['Stock'].tolist()
symbols = [e[0:] for e in symbols]
signaltickerlist = []


#Check all for all tickers if VCP pattern exists
for ticker in symbols:
    dataall = pdr.get_data_yahoo(ticker, start= start, end= now)
    dataall['Date'] = dataall.index
    dataall.fillna(0)  # dropping the NaN values

    volume = dataall['Volume'][-1]
    currentPrice = dataall['Adj Close'][-1]
    try:
        mktcap = int(pdr.get_quote_yahoo(ticker)['marketCap'])
    except IndexError:
        mktcap = 0
    except KeyError:
        mktcap = 0

    cry = VCP_finder(dataall, delta=delta, peakupperbound=peakupperbound, peaklowerbound=peaklowerbound, ddleniency=ddleniency)

    if cry[0] == True:
        signaltickerlist.append([ticker, volume, mktcap, currentPrice])
        print(ticker)
        print(cry[1])

signaltickerlist = (sorted(signaltickerlist))
print(signaltickerlist)

#=============================================

wb = load_workbook(filename=Output_file)
ws = wb['Sheet1']

for i in signaltickerlist:
    ws.append([now, i[0], i[1], i[2], i[3]])

wb.save(filename=Output_file)
